# -*- coding: utf-8 -*-
"""
校准源 xlsx 的「活动直播」打标 —— 源表把「活动直播」当成了广义活动/ Campaign 兜底桶
（展会/颁奖/抽奖/生日都标成了活动直播），导致该标签与真实直播内容严重不符。

本脚本对源表「内容主题」列做**双向、幂等**校准：
  1) 假阳性：当前含「活动直播」但正文无直播信号 -> 去掉「活动直播」
     · 若是组合主题（如「促销活动、活动直播」），仅移除活动直播，保留其余真实主题；
     · 若是纯「活动直播」且无其他主题，按正文重归类到最接近的真实主题
       （抽奖/促销 -> 促销活动；展会/颁奖/发布 -> 资讯转发；生日/周年/祝福 -> 品牌调性）。
  2) 假阴性：当前不含「活动直播」但正文有强直播信号（直播/直播间/livestream/goes live/
     live shopping/live Q&A/broadcast 等）-> 补打「活动直播」。

幂等性：校准后再跑一次，已修正的行不会被二次改动（假阳性行已无活动直播且非直播正文，
不会误加；已补打行同时含活动直播且正文仍匹配直播，保持）。

用法：
  python3 scripts/calibrate_source.py            # 就地写回源 xlsx
  python3 scripts/calibrate_source.py --dry-run  # 只打印将要做的改动，不写文件

注意：本脚本会**修改原始 xlsx**（即工作路径中的「原始数据」）。运行前请先备份。
"""
import re
import argparse
import openpyxl

XLSX = "/Users/fsw/Downloads/GTM社媒数据_打标全表_爆款版.xlsx"
SHEET = "Sheet1"
TOPIC_COL = "内容主题"          # 要校准的列
TEXT_COLS = ["内容文本", "内容文本翻译"]

# —— 直播信号识别 ——
# 难点：英文营销里 "Sale goes live / giveaway is live / website is live" 的 live 是「上线/开售」，
# 不是直播。因此分两级：
#   强信号：中文直播词、livestream、以及「事件本身就是直播」的复合名词（Live Show/Shopping/Q&A…）、live @链接
#   弱信号：goes/we're/i'm + live 等广播谓语（需排除「上线词 + live」的误判）
LIVE_STRONG = re.compile(
    r"(直播|直播间|开播|连麦|直播中)"                                  # 中文直播词
    r"|livestream|live[- ]?stream|streaming|\bbroadcast\b"             # 英文直播复合词/直播
    r"|\blive\s+(show|shows|shopping|q&?a|event|performance|session|demo|broadcast|chat|takeover|auction|sale\s+host)\b"  # 直播事件复合名词
    r"|\blive\s+@|\blive\s+https?://",                                 # live @某人 / live 链接
    re.IGNORECASE,
)
LIVE_WEAK = re.compile(
    r"\b(goings?|goes|went|come|catch|watch|join|tune in|chat|shop|stream|i'?m)\s+live\b"  # 动词 + live（广播）
    r"|\bwe'?re\s+live\b|\bwe\s+are\s+live\b"                          # 品牌/主播正在直播
    r"|\bis\s+live\s+now\b|\bare\s+live\s+now\b|\blive\s+now\b",       # live now（广播中）
    re.IGNORECASE,
)
# 上线词：这些词 + live = 「上线/开售」，不是直播
LAUNCH_NOUNS = (r"(sale|giveaway|deals?|deal|website|site|cart|challenge|collection|drop|"
                r"shop|store|page|contest|promo|promotion|offer|pre[- ]?order|preorder|"
                r"product|launch|update|version|bundle)")
LAUNCH_RE = re.compile(r"\b" + LAUNCH_NOUNS + r"\s+(is|are|was|were|goes|going|went|be)\s+live\b",
                       re.IGNORECASE)


def is_live_content(text):
    # 归一化 HTML 实体（源文本常见 &amp; 即 &，会让 "Q&A" 失配）
    t = (text or "").replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
    if LIVE_STRONG.search(t):
        return True
    m = LIVE_WEAK.search(t)
    if not m:
        return False
    # 弱信号：检查匹配位置附近是否是「上线词 + live」的误判
    start = m.start()
    window = t[max(0, start - 40):start + 15]
    if LAUNCH_RE.search(window):
        return False
    return True

# —— 纯「活动直播」误标后的重归类规则（按顺序命中）——
PROMO_KW = ["抽奖", "giveaway", "sweepstakes", "折扣", "discount", "sale", "优惠",
            "coupon", "促销", "秒杀", "deal", "满减", "赠", "winner", "获奖", "获奖名单", "大赛", "competition"]
NEWS_KW = ["展会", "expo", "exhibition", "booth", "展位", "tradeshow", "convention",
           "颁奖", "论坛", "forum", "峰会", "summit", "回顾", "年度", "发布", "launch"]
BRAND_KW = ["生日", "birthday", "周年", "anniversary", "节日", "祝福", "感谢", "品牌"]


def split_multi(v):
    if not v:
        return []
    return [x.strip() for x in re.split(r"[、,，/]", str(v)) if x.strip()]


def reassign_pure(rec_text):
    t = (rec_text or "").lower()
    if any(k in t for k in PROMO_KW):
        return "促销活动"
    if any(k in t for k in NEWS_KW):
        return "资讯转发"
    if any(k in t for k in BRAND_KW):
        return "品牌调性"
    return "品牌调性"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="只报告不改文件")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]
    first = next(ws.iter_rows(min_row=1, max_row=1))
    header = [str(c.value) if c.value is not None else "" for c in first]
    ci = header.index(TOPIC_COL)
    text_idx = [header.index(c) for c in TEXT_COLS if c in header]

    removed = 0          # 去掉活动直播（假阳性）
    reassigned = {}      # 纯活动直播重归类分布
    added = 0            # 补打活动直播（假阴性）
    kept = 0             # 已正确保留

    for row in ws.iter_rows(min_row=2):
        cell = row[ci]
        raw = cell.value
        if raw is None:
            continue
        tokens = split_multi(raw)
        has_live_topic = "活动直播" in tokens
        text = " ".join(str(row[ti].value or "") for ti in text_idx)
        live = is_live_content(text)

        if has_live_topic and not live:
            # 假阳性：移除活动直播
            new_tokens = [t for t in tokens if t != "活动直播"]
            if not new_tokens:
                # 纯活动直播 -> 重归类
                new_t = reassign_pure(text)
                new_tokens = [new_t]
                reassigned[new_t] = reassigned.get(new_t, 0) + 1
            cell.value = "、".join(new_tokens)
            removed += 1
        elif not has_live_topic and live:
            # 假阴性：补打活动直播
            tokens.append("活动直播")
            cell.value = "、".join(tokens)
            added += 1
        elif has_live_topic and live:
            kept += 1

    print(f"[校准报告] 源表: {XLSX}")
    print(f"  去掉「活动直播」(假阳性, 非直播内容): {removed}")
    print(f"    其中纯活动直播重归类分布: {reassigned}")
    print(f"  补打「活动直播」(假阴性, 真直播漏标): {added}")
    print(f"  已正确保留(正文确为直播): {kept}")

    if args.dry_run:
        print("\n[DRY-RUN] 未写回文件。")
        return

    wb.save(XLSX)
    print("\n已写回源 xlsx。")


if __name__ == "__main__":
    main()
